class ExportExcel:
    def __init__(self, bulletins):
        self.bulletins = bulletins

    def generer(self, fichier=None):
        if fichier:
            self._generer_excel(fichier)
            return fichier
        return self.generer_etat_salarial()

    def generer_etat_salarial(self):
        total_gains = sum(b.total_gains() for b in self.bulletins)
        total_retenues = sum(b.total_retenues() for b in self.bulletins)
        total_net = sum(b.net_a_payer() for b in self.bulletins)

        return {
            "periode": self.bulletins[0].periode if self.bulletins else "",
            "nb_employes": len(self.bulletins),
            "total_gains": total_gains,
            "total_retenues": total_retenues,
            "masse_salariale": total_net,
            "bulletins": [
                {
                    "employe_id": b.employe_id,
                    "periode": b.periode,
                    "brut": float(b.total_gains()),
                    "net": float(b.net_a_payer()),
                }
                for b in self.bulletins
            ],
        }

    def _generer_excel(self, fichier):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Bulletins de paie"

        headers = ["Employé", "Période", "Salaire brut", "Total retenues", "Net à payer"]
        ws.append(headers)

        for col in range(1, 6):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        for bulletin in self.bulletins:
            ws.append([
                bulletin.employe_id,
                bulletin.periode,
                float(bulletin.total_gains()),
                float(bulletin.total_retenues()),
                float(bulletin.net_a_payer()),
            ])

        wb.save(fichier)


def generer_bulletins_excel(bulletins, chemin):
    export = ExportExcel(bulletins)
    return export.generer(chemin)
